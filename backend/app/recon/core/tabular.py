"""Read .csv / .xlsx into a raw grid.

This module exists instead of app.services.uploading.extract_and_clean, which the
rest of the app uses. Two reasons, both disqualifying:

1. extract_and_clean flattens a workbook to a TEXT blob and _clean_text drops
   duplicate lines (uploading.py:67-73). Two byte-identical purchase-register
   rows are exactly what spec section 2.5 requires us to detect and flag as
   Duplicate Purchase Register Entry. That path would delete the evidence before
   the engine ever saw it.
2. Round-tripping a structured file through text and back through csv.reader is
   lossy for no gain: it turns real datetime cells into ambiguous strings.

So: openpyxl for .xlsx (native datetimes survive, which removes date-order
ambiguity for those cells entirely) and the stdlib csv module for .csv (ragged
rows survive, which pandas' column inference would truncate when a banner row is
one cell wide).

There is no dedupe anywhere in this file. Only audited structural rows (blank
spacers, a second header row, and conservative total footers) are omitted.
test_tabular.py locks both behaviours in.
"""
from __future__ import annotations

import csv
import os
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Mapping

from app.recon.core.errors import ParseError
from app.recon.core.normalize import is_valid_gstin, norm_text, to_amount

__all__ = ["OmittedRow", "Table", "read_table", "sheet_names"]

_HEADER_SCAN_ROWS = 40
_MIN_HEADER_SCORE = 3
_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
_HEADER_JOIN = " / "
_DATE_SHAPE_RE = re.compile(r"^\s*\d{1,4}[-/.]\d{1,2}[-/.]\d{1,4}\s*$")
_FOOTER_RE = re.compile(r"^(?:grand\s+total|sub\s*total|total|summary|closing\s+balance)\b", re.I)


@dataclass(frozen=True)
class OmittedRow:
    """A non-transaction row deliberately left out, with an auditable reason."""

    row_no: int
    reason: str
    label: str | None = None


@dataclass
class Table:
    """A source file read into headers plus transaction rows."""

    headers: list[str]
    rows: list[list[Any]]
    header_row: int
    header_end_row: int | None = None
    source_row_numbers: list[int] = field(default_factory=list)
    omitted_rows: list[OmittedRow] = field(default_factory=list)
    sheet_name: str | None = None
    sheet_names: list[str] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)

    def records(self) -> list[dict[str, Any]]:
        """Rows as header->value dicts, positionally aligned.

        Short rows yield None for the missing tail; long rows keep their extra
        cells under generated 'column_<n>' keys rather than losing them.
        """
        out: list[dict[str, Any]] = []
        for row in self.rows:
            record: dict[str, Any] = {}
            for index, header in enumerate(self.headers):
                record[header] = row[index] if index < len(row) else None
            for index in range(len(self.headers), len(row)):
                record[f"column_{index}"] = row[index]
            out.append(record)
        return out


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _looks_like_label(value: Any) -> bool:
    """True for a cell that could be a column heading.

    Headings are words; data is numbers and dates. Counting merely non-empty
    cells is not enough to find a header row — a 3-column data row would outscore
    a 2-column header — so the fallback counts word-ish cells instead.
    """
    if _is_blank(value):
        return False
    if isinstance(value, (int, float, datetime, date)):
        return False
    text = str(value).strip()
    try:
        float(text.replace(",", ""))
    except ValueError:
        return True
    return False  # a numeric string is data, not a heading


def _score_header(cells: list[Any], alias_index: Mapping[str, str] | None) -> int:
    if not alias_index:
        return 0
    return sum(1 for cell in cells if norm_text(cell) in alias_index)


def _looks_like_data(value: Any) -> bool:
    """Use the existing GSTIN/amount regex path when separating headers from data."""
    if _is_blank(value):
        return False
    if isinstance(value, (int, float, datetime, date)):
        return True
    text = str(value).strip()
    if is_valid_gstin(text) or _DATE_SHAPE_RE.match(text):
        return True
    return to_amount(text) is not None


def _find_header_row(grid: list[list[Any]], alias_index: Mapping[str, str] | None) -> int:
    """Pick the header row.

    Government 2B exports and Tally registers both carry several banner/metadata
    rows before the real header, so row 0 is usually wrong.

    Prefer the row that maps to the most known aliases. If nothing scores (an
    unrecognised layout), fall back to the row with the most label-like cells
    rather than to row 0 — a one-cell title banner should never win.

    Ties go to the earliest row: in a real file the header precedes its data.
    """
    best_index, best_score = -1, _MIN_HEADER_SCORE - 1
    for index, cells in enumerate(grid[:_HEADER_SCAN_ROWS]):
        score = _score_header(cells, alias_index)
        if score > best_score:
            best_index, best_score = index, score
    if best_index >= 0:
        return best_index

    best_index, best_labels = 0, -1
    for index, cells in enumerate(grid[:_HEADER_SCAN_ROWS]):
        labels = sum(1 for cell in cells if _looks_like_label(cell))
        if labels > best_labels:
            best_index, best_labels = index, labels
    return best_index


def _is_header_continuation(cells: list[Any], alias_index: Mapping[str, str] | None) -> bool:
    labels = sum(1 for cell in cells if _looks_like_label(cell))
    data = sum(1 for cell in cells if _looks_like_data(cell))
    return labels >= 2 and data == 0 and _score_header(cells, alias_index) > 0


def _find_header_block(
    grid: list[list[Any]], alias_index: Mapping[str, str] | None
) -> tuple[int, int]:
    """Return the inclusive one- or two-row header block.

    Government exports commonly split group labels and leaf labels across two
    adjacent rows. We only extend to an adjacent row that contains known header
    aliases and no GSTIN/date/amount-shaped values.
    """
    selected = _find_header_row(grid, alias_index)
    start = end = selected
    if selected > 0 and _is_header_continuation(grid[selected - 1], alias_index):
        start = selected - 1
    if selected + 1 < len(grid) and _is_header_continuation(grid[selected + 1], alias_index):
        end = selected + 1
    return start, end


def _clean_headers(cells: list[Any]) -> list[str]:
    """Header labels, kept human-readable. Blanks become positional placeholders
    so a nameless column stays addressable in the mapping UI."""
    headers: list[str] = []
    for index, cell in enumerate(cells):
        text = "" if cell is None else str(cell).strip()
        headers.append(text or f"column_{index}")
    return headers


def _flatten_headers(rows: list[list[Any]]) -> list[str]:
    if len(rows) == 1:
        return _clean_headers(rows[0])

    width = max(len(row) for row in rows)
    parents: list[str] = []
    current_parent = ""
    for index in range(width):
        value = rows[0][index] if index < len(rows[0]) else None
        text = "" if value is None else str(value).strip()
        if text:
            current_parent = text
        parents.append(current_parent)

    headers: list[str] = []
    child_row = rows[-1]
    for index in range(width):
        child_value = child_row[index] if index < len(child_row) else None
        child = "" if child_value is None else str(child_value).strip()
        parent = parents[index]
        if parent and child and norm_text(parent) != norm_text(child):
            header = f"{parent}{_HEADER_JOIN}{child}"
        else:
            header = child or parent
        headers.append(header or f"column_{index}")
    return headers


def _field_for_header(header: str, alias_index: Mapping[str, str] | None) -> str | None:
    if not alias_index:
        return None
    field = alias_index.get(norm_text(header))
    if field:
        return field
    parts = header.split(_HEADER_JOIN)
    return alias_index.get(norm_text(parts[-1])) if len(parts) > 1 else None


def _footer_label(cells: list[Any]) -> str | None:
    for cell in cells:
        if isinstance(cell, str) and _FOOTER_RE.match(cell.strip()):
            return cell.strip()
    return None


def _is_total_footer(
    cells: list[Any], headers: list[str], alias_index: Mapping[str, str] | None
) -> tuple[bool, str | None]:
    """Conservatively classify totals; incomplete/malformed transactions survive."""
    label = _footer_label(cells)
    if not label:
        return False, None
    identity_indexes = [
        index
        for index, header in enumerate(headers)
        if _field_for_header(header, alias_index) in {"supplier_gstin", "doc_type", "doc_no", "doc_date"}
    ]
    if len(identity_indexes) < 2:
        return False, None
    identity_values = [cells[index] for index in identity_indexes if index < len(cells)]
    if any(
        not _is_blank(value) and not (isinstance(value, str) and value.strip() == label)
        for value in identity_values
    ):
        return False, None
    has_amount = any(to_amount(cell) is not None for cell in cells if not _is_blank(cell))
    return has_amount, label


def _read_csv_grid(path: str) -> list[list[Any]]:
    last_error: Exception | None = None
    for encoding in _ENCODINGS:
        try:
            with open(path, "r", encoding=encoding, newline="") as handle:
                sample = handle.read(8192)
                handle.seek(0)
                try:
                    dialect: Any = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
                return [list(row) for row in csv.reader(handle, dialect)]
        except UnicodeDecodeError as exc:
            last_error = exc
        except csv.Error as exc:
            raise ParseError(f"Could not read CSV: {exc}") from exc
    raise ParseError(f"Could not decode CSV in any of {', '.join(_ENCODINGS)}") from last_error


def _read_xlsx_grid(path: str, sheet: str | None) -> tuple[list[list[Any]], str, list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ParseError("openpyxl is required to read .xlsx files") from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ParseError(f"Could not open workbook: {exc}") from exc

    try:
        available = list(workbook.sheetnames)
        if sheet is not None:
            if sheet not in available:
                raise ParseError(f"Sheet {sheet!r} not found. Available: {', '.join(available)}")
            worksheet = workbook[sheet]
        else:
            worksheet = _first_nonempty_sheet(workbook, available)
        grid = [list(row) for row in worksheet.iter_rows(values_only=True)]
        return grid, worksheet.title, available
    finally:
        workbook.close()


def _first_nonempty_sheet(workbook: Any, available: list[str]) -> Any:
    """A GSTR-2B export leads with a 'Read me' sheet. Skip past empties, but never
    guess further than that — the caller gets sheet_names and can offer a picker."""
    for name in available:
        worksheet = workbook[name]
        for row in worksheet.iter_rows(values_only=True):
            if any(not _is_blank(cell) for cell in row):
                return worksheet
    return workbook[available[0]]


def sheet_names(path: str) -> list[str]:
    """Sheet names in a workbook; a single synthetic name for a CSV."""
    if os.path.splitext(path)[1].lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    return []


def read_table(
    path: str,
    *,
    alias_index: Mapping[str, str] | None = None,
    sheet: str | None = None,
) -> Table:
    """Read a .csv/.xlsx into a Table. Never dedupes or aggregates. Structural
    rows are omitted only when they satisfy the audited rules above.

    `alias_index` maps norm_text(header) -> system field and is only used to
    locate the header row past any banner rows.
    """
    extension = os.path.splitext(path)[1].lower()
    if extension in (".xlsx", ".xlsm"):
        grid, sheet_name, available = _read_xlsx_grid(path, sheet)
    elif extension == ".csv":
        grid, sheet_name, available = _read_csv_grid(path), None, []
    else:
        raise ParseError(f"Unsupported file type {extension!r}. Upload a .csv or .xlsx file.")

    if not grid:
        raise ParseError("The file is empty.")

    header_row, header_end_row = _find_header_block(grid, alias_index)
    headers = _flatten_headers(grid[header_row : header_end_row + 1])

    rows: list[list[Any]] = []
    source_row_numbers: list[int] = []
    omitted_rows: list[OmittedRow] = []
    if header_end_row > header_row:
        continuation = " | ".join(
            str(cell).strip() for cell in grid[header_end_row] if not _is_blank(cell)
        )
        omitted_rows.append(OmittedRow(header_end_row + 1, "header_continuation", continuation or None))
    for grid_index, cells in enumerate(grid[header_end_row + 1 :], start=header_end_row + 1):
        if all(_is_blank(cell) for cell in cells):
            omitted_rows.append(OmittedRow(grid_index + 1, "blank_row"))
            continue
        is_footer, label = _is_total_footer(cells, headers, alias_index)
        if is_footer:
            omitted_rows.append(OmittedRow(grid_index + 1, "total_footer", label))
            continue
        rows.append(list(cells))
        source_row_numbers.append(grid_index + 1)

    return Table(
        headers=headers,
        rows=rows,
        header_row=header_row,
        header_end_row=header_end_row,
        source_row_numbers=source_row_numbers,
        omitted_rows=omitted_rows,
        sheet_name=sheet_name,
        sheet_names=available,
    )
