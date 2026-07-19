"""Excel download of a reconciliation run (spec section 7).

The contemporary list is shown in the app and exported to Excel ONLY on request,
via a Download Excel button. Status codes are backend sort keys and never appear
in the workbook (spec 5, 6.2) — only the CA-facing status label does.

openpyxl is already a declared dependency (used by the GSTR-1 writer). We build a
fresh workbook in memory and hand back bytes; nothing is written to disk here.
"""
from __future__ import annotations

import io
from typing import Any

# Formula-injection guard: a leading =, +, -, @ makes Excel treat text as a
# formula. Prefix such cells with a single quote so they render as literal text.
_FORMULA_LEADERS = ("=", "+", "-", "@")

_COLUMNS = [
    ("supplier_gstin", "Supplier GSTIN"),
    ("supplier_name", "Trade / Legal Name"),
    ("doc_type", "Document Type"),
    ("doc_no", "Document Number"),
    ("doc_date", "Document Date"),
    ("pr_invoice", "PR Invoice Value"),
    ("cp_invoice", "2B/IMS Invoice Value"),
    ("pr_taxable", "PR Taxable Value"),
    ("cp_taxable", "2B/IMS Taxable Value"),
    ("pr_tax", "PR Combined Tax"),
    ("cp_tax", "2B/IMS Combined Tax"),
    ("diff_invoice", "Invoice Difference"),
    ("diff_pct_display", "Difference %"),
    ("status_label", "Status"),
    ("reason", "Reason"),
]

_IMS_COLUMN = ("ims_status", "IMS Status")
_IMS_INWARD_ACTION_COLUMN = ("ims_action", "IMS Action Status")

_EXCLUDED_COLUMNS = [
    ("side", "Source"),
    ("supplier_gstin", "Supplier GSTIN"),
    ("supplier_name", "Trade / Legal Name"),
    ("doc_type", "Document Type"),
    ("doc_no", "Document Number"),
    ("doc_date", "Document Date"),
    ("invoice", "Invoice Value"),
    ("reason", "Held Out Because"),
]

_MODULE_TITLES = {
    "gstr2b": "GSTR-2B Reconciliation",
    "ims_outward": "IMS Outward Reconciliation",
    "ims_inward": "IMS Inward",
}


def build_workbook(
    results: list[dict[str, Any]],
    excluded: list[dict[str, Any]],
    *,
    module: str,
    period: str | None = None,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    columns = list(_COLUMNS)
    if module == "ims_inward":
        columns.insert(4, _IMS_INWARD_ACTION_COLUMN)
    elif module == "ims_outward":
        columns.insert(4, _IMS_COLUMN)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reconciliation"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F5132")

    for col, (_, label) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill

    for r, result in enumerate(results, start=2):
        row = _display_row(result)
        for col, (key, _) in enumerate(columns, start=1):
            sheet.cell(row=r, column=col, value=_safe(row.get(key)))

    _autosize(sheet, columns)

    if excluded:
        ex_sheet = workbook.create_sheet("Excluded")
        for col, (_, label) in enumerate(_EXCLUDED_COLUMNS, start=1):
            cell = ex_sheet.cell(row=1, column=col, value=label)
            cell.font = header_font
            cell.fill = header_fill
        for r, item in enumerate(excluded, start=2):
            for col, (key, _) in enumerate(_EXCLUDED_COLUMNS, start=1):
                value = _reason_label(item[key]) if key == "reason" else item.get(key)
                ex_sheet.cell(row=r, column=col, value=_safe(value))
        _autosize(ex_sheet, _EXCLUDED_COLUMNS)

    buffer = io.BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def _display_row(result: dict[str, Any]) -> dict[str, Any]:
    diff_pct = result.get("diff_pct")
    return {
        **result,
        "diff_pct_display": None if diff_pct is None else round(diff_pct * 100, 2),
    }


def _reason_label(reason: str) -> str:
    return {"out_of_period": "Dated outside the selected period"}.get(reason, reason)


def _safe(value: Any) -> Any:
    """Neutralize spreadsheet formula injection from untrusted text cells."""
    if isinstance(value, str) and value[:1] in _FORMULA_LEADERS:
        return "'" + value
    return value


def _autosize(sheet, columns: list[tuple[str, str]]) -> None:
    from openpyxl.utils import get_column_letter

    for col, (_, label) in enumerate(columns, start=1):
        width = max(len(label) + 2, 12)
        sheet.column_dimensions[get_column_letter(col)].width = min(width, 40)


def filename_for(module: str, period: str | None) -> str:
    base = _MODULE_TITLES.get(module, "Reconciliation").replace(" ", "_")
    suffix = f"_{period}" if period else ""
    return f"{base}{suffix}.xlsx"
